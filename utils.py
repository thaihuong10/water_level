from openpyxl import load_workbook
from datetime import datetime

def read_excel_openpyxl(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    data = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and "Thời gian" in str(row[1]) and "Mực nước" in str(row[5]):
                header_found = True
            continue
        if row and row[1] and row[5]:
            try:
                ts = datetime.strptime(str(row[1]), "%d/%m/%Y %H:%M:%S") if isinstance(row[1], str) else row[1]
                data.append({
                    "timestamp": ts,
                    "water_level": float(row[5])
                })
            except:
                continue

    return data

def get_daily_extremes(data, date_str):
    try:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        raise ValueError("Sai định dạng ngày.")

    filtered = [d for d in data if d["timestamp"].date() == selected_date]
    if not filtered:
        return [], []

    max_val = max(d["water_level"] for d in filtered)
    min_val = min(d["water_level"] for d in filtered)

    max_entries = [d for d in filtered if d["water_level"] == max_val]
    min_entries = [d for d in filtered if d["water_level"] == min_val]

    return max_entries, min_entries
